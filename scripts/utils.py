import os
import pandas as pd
from snowflake_connector import create_snowflake_connection
import xlwings as xw
from datetime import datetime, timedelta
import shutil
import win32com.client as win32
import smtplib
from email.message import EmailMessage
from email.utils import make_msgid
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
import time
import openpyxl
from datetime import datetime
import os
import re
import shutil
import glob

def get_previous_month():
    today = datetime.today()
    first = today.replace(day=1)
    prev_month = first.replace(month=first.month - 1 if first.month > 1 else 12, 
                               year=first.year if first.month > 1 else first.year - 1)
    return prev_month.strftime("%B")


# Function to read SQL queries from files
def read_query(file_path):
    try:
        with open(file_path, 'r') as file:
            query = file.read()
        return query
    except FileNotFoundError:
        print(f"Error: File not found - {file_path}")
        return None

# Function to fetch data from Snowflake and return a DataFrame
def fetch_data(query):
    try:
        conn = create_snowflake_connection()
    except Exception as e:
        print(f"Error connecting to Snowflake: {e}")
        return None

    try:
        cur = conn.cursor()
        cur.execute(query)
        results = cur.fetchall()

        if not results:
            print("Warning: No data returned from query.")
            return pd.DataFrame()

        df = pd.DataFrame(results, columns=[desc[0] for desc in cur.description])
    except Exception as e:
        print(f"Error executing query: {e}")
        return pd.DataFrame() 
    finally:
        cur.close()
        conn.close()

    return df


# Function to read and replace placeholder in SQL query
def get_personalized_query(query_template_file, name):
    with open(query_template_file, 'r') as file:
        query_template = file.read()
    return query_template.replace("{SDDR_NAME}", name)

# Function to save data into an Excel file with multiple sheets
def save_to_excel(excel_file_path, query_files, individual):
    os.makedirs(os.path.dirname(excel_file_path), exist_ok=True)
    # Initialize an Excel writer
    try:
        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            for sheet_name, query_file in query_files.items():
                if individual == True:
                    query = query_file
                else:
                    query = read_query(query_file)
                if query:
                    df = fetch_data(query)
                    if df.empty:
                        # If the data is empty, create an empty DataFrame with the correct columns
                        placeholder_df = pd.DataFrame(columns=df.columns)
                        placeholder_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        print(f"Warning: No data for {sheet_name}. Empty table with headers saved.")
                    else:
                        # If there is data, write it as is
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    print(f"Error: Could not read query file for {sheet_name}. Skipping.")
    except Exception as e:
        print(f"Error saving data to Excel: {e}")
    else:
        print(f"Data saved to '{excel_file_path}' successfully!")
        

def copy_table(source_ws, target_ws, source_range, target_start_cell):
    # Select the source range and target range
    source_cells = source_ws.range((source_range[0], source_range[1]), (source_range[2], source_range[3]))
    target_start = target_ws.range(target_start_cell)
    
    # Copy the table from source to target
    source_cells.copy(target_start)

# Function to replace table
def replace_table(source_file, source_sheet_name, source_table_range, target_file, target_sheet_name, target_start_cell):
    previous_month = get_previous_month()
    max_retries = 3  # Maximum number of retries
    attempt = 0

    while attempt < max_retries:
        app = xw.App(visible=False)
        
        try:
            source_wb = app.books.open(source_file)
            target_wb = app.books.open(target_file)
            
            try:
                source_ws = source_wb.sheets[source_sheet_name]
                target_ws = target_wb.sheets[target_sheet_name]

                # Copy the table from source to target
                copy_table(source_ws, target_ws, source_table_range, target_start_cell)

                # Replace the content in cell I7 with the previous month
                target_ws.range('I7').value = previous_month

                # Save the changes to the target workbook
                target_wb.save(target_file)
                print(f"Table updated in '{target_sheet_name}' of '{target_file}' successfully.")
                return  # Exit the function if operation is successful

            finally:
                # Close the workbooks after operations
                source_wb.close()
                target_wb.close()

        except Exception as e:
            if "Cannot access" in str(e) and "xlmain11.chm" in str(e):
                print("Encountered known file access error, but operations completed successfully.")
                return  # Exit the function if known error is handled

            print(f"Unexpected error on attempt {attempt + 1}: {e}")
            attempt += 1
            time.sleep(2)

        finally:
            app.quit()

    # If all retries fail, return -1
    print("All retries failed. Operation could not be completed.")
    return -1

# Function to remove trailing zeros
def remove_trailing_zeros(file_path, sheet_name, column_name):
    app = xw.App(visible=False)
    
    try:
        # Open the workbook and select the sheet
        wb = app.books.open(file_path)
        sheet = wb.sheets[sheet_name]

        # Find the column index for the specified column name
        col_range = sheet.range(1, 1).expand('right').value
        stat_col = None
        for i, col_name in enumerate(col_range):
            if col_name == column_name:
                stat_col = i + 1
                break

        if not stat_col:
            print(f"Column '{column_name}' not found!")
            return

        # Get the last row with data in the column
        last_row = sheet.range(1, stat_col).end('down').row

        # Iterate over each cell in the column to apply formatting
        for row in range(2, last_row + 1):
            cell = sheet.range((row, stat_col))
            value = cell.value
            
            # Convert text-based numbers to float
            if isinstance(value, str):
                try:
                    value = float(value)
                except ValueError:
                    continue

            # Format value as an Excel text formula with thousands separator
            if isinstance(value, (int, float)):
                if value.is_integer():
                    # Format integer as ="#,###"
                    cell.value = f'="{int(value):,}"'
                else:
                    # Format float with up to 6 decimals and remove unnecessary trailing zeros
                    formatted_value = f"{value:,.6f}".rstrip('0').rstrip('.')
                    cell.value = f'="{formatted_value}"'

        wb.save()
        print(f"Values formatted as text formulas with commas in column '{column_name}' of sheet '{sheet_name}' successfully.")
    except Exception as e:
        print(f"Error while formatting column '{column_name}': {e}")
    finally:
        wb.close()
        app.quit()
        
def create_folder_structure(base_path, folder_name):
    # Define the "Hyperlink" folder path
    folder_with_name = os.path.join(base_path, folder_name)
    
    # Get current date and derive previous and two-months-ago dates
    today = datetime.today()
    first_day_of_current_month = today.replace(day=1)
    previous_month_date = first_day_of_current_month - timedelta(days=1)

    # Format folders as "MM.YYYY"
    previous_month_folder = f"{previous_month_date.month}.{previous_month_date.year}"

    # Create "Hyperlink" folder if it doesn't exist
    if not os.path.exists(folder_with_name):
        os.makedirs(folder_with_name)
    
    # Create the folder for the previous month
    previous_month_path = os.path.join(folder_with_name, previous_month_folder)
    if not os.path.exists(previous_month_path):
        os.makedirs(previous_month_path)


    print(f"Folder structure verified. Previous month folder is: {previous_month_path}")
    return previous_month_path


def close_existing_workbook(app, workbook_name):
    """Close workbook if it's already open."""
    for wb in app.books:
        if wb.name == workbook_name:
            wb.close(SaveChanges=False)
            print(f"Closed existing workbook: {workbook_name}")

def save_selection_as_pdf(excel_file, sheet_name, selection_range, output_pdf):
    app = xw.App(visible=False)
    app.display_alerts = False  # Disable pop-ups in Excel for smoother handling
    
    # Get the name of the workbook file from the path
    workbook_name = os.path.basename(excel_file)
    
    try:
        # Close the workbook if already open
        close_existing_workbook(app, workbook_name)
        # Open workbook and select the specific sheet
        workbook = app.books.open(excel_file)
        sheet = workbook.sheets[sheet_name]
        # Select the specified range
        selected_range = sheet.range(selection_range)

        # Ensure the output directory exists
        output_dir = os.path.dirname(output_pdf)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        # Use a temporary PDF file location
        temp_pdf_path = os.path.abspath(os.path.join(output_dir, "temp_output.pdf"))

        # Export selection as PDF
        selected_range.api.ExportAsFixedFormat(
            Type=0,  # PDF format
            Filename=temp_pdf_path,
            Quality=0,  # Standard quality
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            From=1,  # Start page
            To=1,    # End page
            OpenAfterPublish=False
        )
        # Ensure the temp PDF file was created
        if os.path.exists(temp_pdf_path):
            # Rename temp file to the desired output
            os.replace(temp_pdf_path, output_pdf)
            print(f"PDF saved successfully to '{output_pdf}'.")
        else:
            print(f"Error: Temporary PDF file was not created at '{temp_pdf_path}'.")

    except Exception as e:
        print(f"An error occurred while saving the PDF: {e}")

    finally:
        # Ensure workbook is closed
        if 'workbook' in locals() and workbook:
            workbook.close()
        app.quit()


# def send_email(subject, body, sender_email, sender_password, recipients, attachment_path, SMTP_SERVER, SMTP_PORT):
#     # Create a multipart email
#     msg = MIMEMultipart()
#     msg['From'] = sender_email
#     msg['To'] = ', '.join(recipients)
#     msg['Subject'] = subject

#     # Attach the body
#     msg.attach(MIMEText(body, 'plain'))

#     # Attach the PDF file
#     with open(attachment_path, 'rb') as f:
#         part = MIMEApplication(f.read(), _subtype='pdf')
#         part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(attachment_path))
#         msg.attach(part)

#     # Send the email
#     try:
#         with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
#             server.starttls()  # Secure the connection
#             server.login(sender_email, sender_password)
#             server.sendmail(sender_email, recipients, msg.as_string())
#             print(f"Email sent successfully to {', '.join(recipients)}")
#     except Exception as e:
#         print(f"An error occurred: {e}")


def send_email(subject, body, sender_email, sender_password, recipients, attachment_path, SMTP_SERVER='smtp.office365.com', SMTP_PORT=587):
    # Create a multipart email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ', '.join(recipients)
    msg['Subject'] = subject

    # Attach the email body
    msg.attach(MIMEText(body, 'plain'))

    # Attach the PDF file
    if attachment_path and os.path.isfile(attachment_path):
        with open(attachment_path, 'rb') as f:
            part = MIMEApplication(f.read(), _subtype='pdf')
            part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(attachment_path))
            msg.attach(part)

    # Send the email
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()  # Secure the connection
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipients, msg.as_string())
            print(f"Email sent successfully to {', '.join(recipients)}")
    except Exception as e:
        print(f"An error occurred while sending the email: {e}")        

def get_previous_month_year():
    today = datetime.today()
    first = today.replace(day=1)
    prev_month = first.replace(month=first.month - 1 if first.month > 1 else 12, 
                               year=first.year if first.month > 1 else first.year - 1)
    return prev_month.strftime("%b"), prev_month.year  # returns abbreviated month name and year

def update_sheet_content(original_sheet, source_sheet):
    # Determine the range of data in the source sheet
    max_row = source_sheet.max_row
    max_col = source_sheet.max_column

    # Clear contents in the original sheet starting from A5 and beyond the source sheet's range
    for row in original_sheet.iter_rows(min_row=5, min_col=1, max_row=max_row + 1000, max_col=max_col+4):
        for cell in row:
            cell.value = None

    # Copy contents from source sheet (starting A2) to original sheet (starting A6)
    for src_row, orig_row in zip(source_sheet.iter_rows(min_row=1, min_col=1, max_row=max_row, max_col=max_col+4),
                                 original_sheet.iter_rows(min_row=5, min_col=1, max_row=max_row + 1000, max_col=max_col+4)):
        for src_cell, orig_cell in zip(src_row, orig_row):
            orig_cell.value = src_cell.value


def update_date_text(cell, prev_month, prev_year):
    # Find and replace date pattern (e.g., "Sept 2024" with "Oct 2024")
    pattern = r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sept|Oct|Nov|Dec) \d{4}"
    if isinstance(cell.value, str):
        cell.value = re.sub(pattern, f"{prev_month} {prev_year}", cell.value)

def update_source_date_text(cell):
    # Update the date in A2 with today’s date in "M.D.YYYY" format
    if isinstance(cell.value, str):
        today = datetime.today().strftime("%m.%d.%Y")
        pattern = r"\b\d{1,2}\.\d{1,2}\.\d{4}\b"  # Match date patterns like "9.30.2024"
        cell.value = re.sub(pattern, today, cell.value)

def update_workbook(original_file_path, source_file_path):
    # Load the original and source workbooks
    original_wb = openpyxl.load_workbook(original_file_path)
    source_wb = openpyxl.load_workbook(source_file_path)

    # Get the previous month and year
    prev_month, prev_year = get_previous_month_year()

    for sheet_name in original_wb.sheetnames:
        original_sheet = original_wb[sheet_name]
        
        # Update A1 in the original sheet with the previous month and year
        update_date_text(original_sheet["A1"], prev_month, prev_year)
        
        # Update A2 in the original sheet with today’s date
        update_source_date_text(original_sheet["A2"])

        # Check if this sheet exists in the source workbook and update contents
        if sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[sheet_name]
            update_sheet_content(original_sheet, source_sheet)
        else:
            print(f"Sheet '{sheet_name}' not found in source file. Skipping...")

    # Save and close workbooks
    original_wb.save(original_file_path)
    original_wb.close()
    source_wb.close()


    print(f"Workbook '{original_file_path}' updated successfully.")